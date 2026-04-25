"""
Servidor de estadísticas, registro de aliados y panel de administración.
Corre en puerto 5055.
"""

import os
import io
import uuid
import sqlite3
from datetime import datetime
from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
import database as db
from config import DB_PATH

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__, static_folder="uploads")
CORS(app)

ADMIN_PASSWORD = "RedCol2026"


# ─── Helpers ────────────────────────────────────────────────────────────────

def _save_upload(file_obj, prefix="img"):
    if not file_obj or file_obj.filename == "":
        return None
    ext = os.path.splitext(file_obj.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".webp"):
        return None
    filename = f"{prefix}_{uuid.uuid4().hex[:10]}{ext}"
    path = os.path.join(UPLOAD_DIR, filename)
    file_obj.save(path)
    return filename


def _raw_query(sql, params=()):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _raw_scalar(sql, params=()):
    conn = sqlite3.connect(DB_PATH)
    val = conn.execute(sql, params).fetchone()[0]
    conn.close()
    return val


# ─── Stats API (landing page) ──────────────────────────────────────────────

@app.route("/api/stats")
def stats():
    db.init_db()
    leaderboard = db.get_leaderboard(20)
    total_users = _raw_scalar("SELECT COUNT(*) FROM users")
    total_completions = _raw_scalar("SELECT COUNT(*) FROM task_completions WHERE status='approved'")
    total_points = _raw_scalar("SELECT COALESCE(SUM(points_awarded),0) FROM task_completions WHERE status='approved'")
    total_redemptions = _raw_scalar("SELECT COUNT(*) FROM redemptions")

    board = [{"rank": i+1, "name": u["full_name"] or u["username"] or f"Usuario {u['user_id']}", "points": u["points"]}
             for i, u in enumerate(leaderboard)]

    return jsonify({
        "leaderboard": board,
        "summary": {"total_users": total_users, "total_completions": total_completions,
                     "total_points_issued": total_points, "total_redemptions": total_redemptions}
    })


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


# ─── Registro de empresas aliadas (público) ─────────────────────────────────

@app.route("/registro")
def registro_form():
    return send_from_directory(TEMPLATES_DIR, "registro_aliado.html")


@app.route("/api/registro", methods=["POST"])
def registrar_aliado():
    db.init_db()
    try:
        business_name = request.form.get("business_name", "").strip()
        if not business_name:
            return jsonify({"ok": False, "error": "El nombre de la empresa es obligatorio"}), 400

        photo_path = _save_upload(request.files.get("business_photo"), "biz")
        telegram_user = request.form.get("telegram_user", "").strip().lstrip("@")
        try:
            tasks_per_week = int(request.form.get("tasks_per_week", 7))
        except ValueError:
            tasks_per_week = 7
        ally_id = db.create_ally(
            business_name,
            request.form.get("owner_name", "").strip(),
            request.form.get("phone", "").strip(),
            request.form.get("email", "").strip(),
            request.form.get("location", "").strip(),
            request.form.get("city", "").strip(),
            request.form.get("description", "").strip(),
            photo_path,
            request.form.get("instagram", "").strip(),
            request.form.get("facebook", "").strip(),
            request.form.get("website", "").strip(),
            telegram_user,
            tasks_per_week,
        )

        for idx in range(10):
            prod_name = request.form.get(f"product_name_{idx}", "").strip()
            if not prod_name:
                break
            prod_photo = _save_upload(request.files.get(f"product_photo_{idx}"), f"prod{idx}")
            db.add_ally_product(ally_id, prod_name,
                                request.form.get(f"product_desc_{idx}", "").strip(),
                                request.form.get(f"product_price_{idx}", "").strip(),
                                prod_photo)

        return jsonify({"ok": True, "ally_id": ally_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─── Agregar productos (link público para empresas registradas) ─────────────

@app.route("/productos/<int:ally_id>")
def productos_form(ally_id):
    return send_from_directory(TEMPLATES_DIR, "agregar_productos.html")


@app.route("/api/productos/<int:ally_id>", methods=["GET"])
def get_ally_info(ally_id):
    db.init_db()
    ally = db.get_ally(ally_id)
    if not ally:
        return jsonify({"ok": False, "error": "Empresa no encontrada"}), 404
    products = db.get_ally_products(ally_id)
    return jsonify({
        "ok": True,
        "ally": {"id": ally["id"], "business_name": ally["business_name"]},
        "products": [{"id": p["id"], "name": p["name"], "description": p["description"],
                       "price": p["price"],
                       "photo": f"/uploads/{p['photo_path']}" if p["photo_path"] else None}
                      for p in products]
    })


@app.route("/api/productos/<int:ally_id>", methods=["POST"])
def add_products(ally_id):
    db.init_db()
    ally = db.get_ally(ally_id)
    if not ally:
        return jsonify({"ok": False, "error": "Empresa no encontrada"}), 404
    try:
        added = 0
        for idx in range(10):
            prod_name = request.form.get(f"product_name_{idx}", "").strip()
            if not prod_name:
                break
            prod_photo = _save_upload(request.files.get(f"product_photo_{idx}"), f"prod{idx}")
            db.add_ally_product(ally_id, prod_name,
                                request.form.get(f"product_desc_{idx}", "").strip(),
                                request.form.get(f"product_price_{idx}", "").strip(),
                                prod_photo)
            added += 1
        return jsonify({"ok": True, "added": added})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─── Panel de Administración ───────────────────────────────────────────────

@app.route("/admin")
def admin_panel():
    return send_from_directory(TEMPLATES_DIR, "admin_panel.html")


@app.route("/api/admin/dashboard")
def admin_dashboard():
    db.init_db()
    total_users = _raw_scalar("SELECT COUNT(*) FROM users")
    total_points = _raw_scalar("SELECT COALESCE(SUM(points),0) FROM users")
    total_tasks = _raw_scalar("SELECT COUNT(*) FROM tasks WHERE is_active=1")
    total_completions = _raw_scalar("SELECT COUNT(*) FROM task_completions WHERE status='approved'")
    pending_completions = _raw_scalar("SELECT COUNT(*) FROM task_completions WHERE status='pending'")
    total_rewards = _raw_scalar("SELECT COUNT(*) FROM rewards WHERE is_active=1")
    total_redemptions = _raw_scalar("SELECT COUNT(*) FROM redemptions")
    total_allies = _raw_scalar("SELECT COUNT(*) FROM allies")
    pending_allies = _raw_scalar("SELECT COUNT(*) FROM allies WHERE status='pending'")
    total_products = _raw_scalar("SELECT COUNT(*) FROM ally_products")

    return jsonify({
        "total_users": total_users, "total_points": total_points,
        "total_tasks": total_tasks, "total_completions": total_completions,
        "pending_completions": pending_completions,
        "total_rewards": total_rewards, "total_redemptions": total_redemptions,
        "total_allies": total_allies, "pending_allies": pending_allies,
        "total_products": total_products,
    })


@app.route("/api/admin/users")
def admin_users():
    db.init_db()
    users = _raw_query("SELECT * FROM users ORDER BY points DESC")
    return jsonify(users)


@app.route("/api/admin/users/<int:uid>", methods=["PUT"])
def admin_update_user(uid):
    db.init_db()
    data = request.json
    if "points" in data:
        conn = sqlite3.connect(DB_PATH)
        conn.execute("UPDATE users SET points = ? WHERE user_id = ?", (int(data["points"]), uid))
        conn.commit()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/users/<int:uid>", methods=["DELETE"])
def admin_delete_user(uid):
    db.init_db()
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM task_completions WHERE user_id = ?", (uid,))
    conn.execute("DELETE FROM redemptions WHERE user_id = ?", (uid,))
    conn.execute("DELETE FROM users WHERE user_id = ?", (uid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/allies")
def admin_allies():
    db.init_db()
    allies = _raw_query("SELECT * FROM allies ORDER BY created_at DESC")
    for a in allies:
        a["products"] = _raw_query("SELECT * FROM ally_products WHERE ally_id = ?", (a["id"],))
    return jsonify(allies)


@app.route("/api/admin/allies/<int:aid>/status", methods=["PUT"])
def admin_ally_status(aid):
    db.init_db()
    data = request.json
    db.update_ally_status(aid, data.get("status", "pending"))
    return jsonify({"ok": True})


@app.route("/api/admin/allies/<int:aid>", methods=["PUT"])
def admin_update_ally(aid):
    db.init_db()
    data = request.json
    db.update_ally(aid, **data)
    return jsonify({"ok": True})


@app.route("/api/admin/allies/<int:aid>", methods=["DELETE"])
def admin_delete_ally(aid):
    db.init_db()
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM ally_products WHERE ally_id = ?", (aid,))
    conn.execute("DELETE FROM allies WHERE id = ?", (aid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/debug/simulate-redeem")
def admin_simulate_redeem():
    """Simula un canje sin enviar mensajes Telegram para diagnosticar."""
    db.init_db()
    import secrets, traceback
    from voucher import generate_voucher

    user_id = int(request.args.get("user_id", 0))
    product_id = int(request.args.get("product_id", 0))

    steps = []
    try:
        steps.append({"step": "get_user", "user_id": user_id})
        user = db.get_user(user_id)
        steps[-1]["result"] = user

        steps.append({"step": "get_product", "product_id": product_id})
        product = db.get_product(product_id)
        steps[-1]["result"] = product

        if not product:
            return jsonify({"ok": False, "steps": steps, "error": "product not found"})
        if not product.get("is_active"):
            return jsonify({"ok": False, "steps": steps, "error": "product not active"})
        if product.get("points_required", 0) <= 0:
            return jsonify({"ok": False, "steps": steps, "error": "points_required <= 0"})
        if not user or user["points"] < product["points_required"]:
            return jsonify({"ok": False, "steps": steps,
                            "error": f"insufficient points: have={user['points'] if user else 0}, need={product['points_required']}"})

        steps.append({"step": "generate_voucher_test"})
        path = generate_voucher(
            user_name=user["full_name"],
            reward_name=product["name"],
            provider=product["provider"],
            points_used=product["points_required"],
            new_balance=user["points"] - product["points_required"],
            voucher_code="SIMULATE"
        )
        steps[-1]["voucher_path"] = path
        steps[-1]["voucher_exists"] = os.path.exists(path) if path else False

        return jsonify({"ok": True, "steps": steps,
                        "would_succeed": True,
                        "would_use_points": product["points_required"]})
    except Exception as e:
        return jsonify({"ok": False, "steps": steps,
                        "error": str(e), "traceback": traceback.format_exc()})


@app.route("/api/admin/debug/test-voucher/<int:product_id>")
def admin_test_voucher(product_id):
    """Genera un voucher de prueba para diagnosticar."""
    db.init_db()
    try:
        from voucher import generate_voucher
        product = db.get_product(product_id)
        if not product:
            return jsonify({"ok": False, "error": "Producto no existe"}), 404
        path = generate_voucher(
            user_name="Test User",
            reward_name=product["name"],
            provider=product["provider"],
            points_used=product["points_required"],
            new_balance=0,
            voucher_code="TEST1234"
        )
        return jsonify({
            "ok": True,
            "path": path,
            "exists": os.path.exists(path) if path else False,
            "size": os.path.getsize(path) if path and os.path.exists(path) else 0,
            "vouchers_dir": os.getenv("VOUCHERS_DIR", "vouchers"),
            "vouchers_dir_exists": os.path.exists(os.getenv("VOUCHERS_DIR", "vouchers"))
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/admin/scheduled")
def admin_scheduled():
    """Lista los scheduled_tasks recientes para debugging."""
    db.init_db()
    rows = _raw_query("""
        SELECT st.id, st.task_id, st.send_at, st.sent, st.message_id,
               st.auto_generated, st.slot_time,
               t.title, t.is_active as task_active
        FROM scheduled_tasks st
        LEFT JOIN tasks t ON t.id = st.task_id
        ORDER BY st.id DESC
        LIMIT 50
    """)
    return jsonify(rows)


@app.route("/api/admin/run-pending-tasks", methods=["POST"])
def admin_run_pending_tasks():
    """Fuerza envío de tareas pendientes via Telegram Bot API directamente."""
    import requests
    from config import BOT_TOKEN, GROUP_CHAT_ID, PROJECT_NAME
    db.init_db()

    pending = _raw_query("""
        SELECT st.id, st.task_id,
               t.title, t.description, t.instructions, t.target_url, t.points_value
        FROM scheduled_tasks st
        JOIN tasks t ON t.id = st.task_id
        WHERE st.sent = 0 AND st.send_at <= datetime('now') AND t.is_active = 1
        ORDER BY st.send_at ASC
    """)

    sent = []
    errors = []
    for task in pending:
        url_line = f"\n\n🔗 Enlace: {task['target_url']}" if task.get("target_url") else ""
        text = (
            f"📢 Nueva Tarea — {PROJECT_NAME}!\n\n"
            f"📌 {task['title']}\n\n"
            f"{task['description']}\n\n"
            f"📋 ¿Cómo completarla?\n"
            f"{task['instructions']}"
            f"{url_line}\n\n"
            f"💰 Vale: {task['points_value']} puntos\n\n"
            f"✅ Cuando termines, envía el screenshot al bot en privado:\n"
            f"👉 @Redcolaboracion_bot"
        )
        try:
            r = requests.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                json={"chat_id": GROUP_CHAT_ID, "text": text},
                timeout=15
            )
            data = r.json()
            if data.get("ok"):
                msg_id = data["result"]["message_id"]
                conn = sqlite3.connect(DB_PATH)
                conn.execute("UPDATE scheduled_tasks SET sent=1, message_id=? WHERE id=?",
                             (msg_id, task["id"]))
                conn.commit()
                conn.close()
                sent.append({"id": task["id"], "title": task["title"], "message_id": msg_id})
            else:
                errors.append({"id": task["id"], "error": data})
        except Exception as e:
            errors.append({"id": task["id"], "error": str(e)})

    return jsonify({"sent": sent, "errors": errors, "total": len(pending)})


@app.route("/api/admin/tasks")
def admin_tasks():
    db.init_db()
    tasks = _raw_query("""
        SELECT t.*,
               COALESCE(a.business_name, '') as ally_name,
               (SELECT COUNT(*) FROM task_completions tc WHERE tc.task_id = t.id AND tc.status='approved') as completed_count,
               (SELECT COUNT(*) FROM task_completions tc WHERE tc.task_id = t.id AND tc.status='pending') as pending_count
        FROM tasks t
        LEFT JOIN allies a ON a.id = t.ally_id
        ORDER BY t.created_at DESC
    """)
    return jsonify(tasks)


@app.route("/api/admin/tasks", methods=["POST"])
def admin_create_task():
    db.init_db()
    data = request.json
    title = data.get("title", "").strip()
    if not title:
        return jsonify({"ok": False, "error": "El titulo es obligatorio"}), 400
    tid = db.create_task(
        title=title,
        description=data.get("description", ""),
        instructions=data.get("instructions", ""),
        target_url=data.get("target_url") or None,
        points_value=int(data.get("points_value", 10)),
        created_by=0,
        ally_id=int(data["ally_id"]) if data.get("ally_id") else None
    )
    return jsonify({"ok": True, "id": tid})


@app.route("/api/admin/tasks/<int:tid>", methods=["PUT"])
def admin_update_task(tid):
    db.init_db()
    data = request.json
    allowed = {"title", "description", "instructions", "target_url", "points_value", "ally_id"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"ok": True})
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [tid]
    conn = sqlite3.connect(DB_PATH)
    conn.execute(f"UPDATE tasks SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/tasks/<int:tid>", methods=["DELETE"])
def admin_delete_task(tid):
    db.init_db()
    conn = sqlite3.connect(DB_PATH)
    conn.execute("UPDATE tasks SET is_active = 0 WHERE id = ?", (tid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/rewards")
def admin_rewards():
    db.init_db()
    return jsonify(_raw_query("SELECT * FROM rewards ORDER BY points_required ASC"))


@app.route("/api/admin/rewards", methods=["POST"])
def admin_create_reward():
    db.init_db()
    data = request.json
    rid = db.create_reward(data["name"], data.get("description", ""), int(data["points_required"]), data.get("provider", ""))
    return jsonify({"ok": True, "id": rid})


@app.route("/api/admin/rewards/<int:rid>", methods=["DELETE"])
def admin_delete_reward(rid):
    db.init_db()
    db.deactivate_reward(rid)
    return jsonify({"ok": True})


@app.route("/api/admin/redemptions")
def admin_redemptions():
    db.init_db()
    rows = _raw_query("""
        SELECT rd.*, r.name as reward_name, r.provider, u.full_name, u.username
        FROM redemptions rd
        JOIN rewards r ON r.id = rd.reward_id
        JOIN users u ON u.user_id = rd.user_id
        ORDER BY rd.redeemed_at DESC
    """)
    return jsonify(rows)


@app.route("/api/admin/completions")
def admin_completions():
    db.init_db()
    rows = _raw_query("""
        SELECT tc.*, t.title, t.points_value, u.full_name, u.username
        FROM task_completions tc
        JOIN tasks t ON t.id = tc.task_id
        JOIN users u ON u.user_id = tc.user_id
        ORDER BY tc.submitted_at DESC
        LIMIT 100
    """)
    return jsonify(rows)


@app.route("/api/admin/products")
def admin_all_products():
    db.init_db()
    rows = _raw_query("""
        SELECT p.*, a.business_name
        FROM ally_products p
        JOIN allies a ON a.id = p.ally_id
        ORDER BY a.business_name, p.name
    """)
    return jsonify(rows)


@app.route("/api/admin/products", methods=["POST"])
def admin_create_product():
    db.init_db()
    data = request.json
    name = data.get("name", "").strip()
    ally_id = data.get("ally_id")
    if not name or not ally_id:
        return jsonify({"ok": False, "error": "Nombre y empresa son obligatorios"}), 400
    pid = db.add_ally_product(
        int(ally_id), name,
        data.get("description", ""),
        data.get("price", ""),
        None,
        int(data.get("points_required", 0))
    )
    return jsonify({"ok": True, "id": pid})


@app.route("/api/admin/products/<int:pid>", methods=["PUT"])
def admin_update_product(pid):
    db.init_db()
    data = request.json
    allowed = {"name", "description", "price", "ally_id", "points_required"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"ok": True})
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [pid]
    conn = sqlite3.connect(DB_PATH)
    conn.execute(f"UPDATE ally_products SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/products/<int:pid>", methods=["DELETE"])
def admin_delete_product(pid):
    db.init_db()
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM ally_products WHERE id = ?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─── Estadisticas avanzadas ────────────────────────────────────────────────

@app.route("/api/admin/stats/by-ally")
def admin_stats_by_ally():
    """Estadísticas agregadas por empresa, opcionalmente filtradas por mes."""
    db.init_db()
    month = request.args.get("month", "")  # "YYYY-MM"

    tc_filter = ""
    red_filter = ""
    params = []
    if month:
        tc_filter = " AND strftime('%Y-%m', tc.submitted_at) = ?"
        red_filter = " AND strftime('%Y-%m', r.redeemed_at) = ?"
        params = [month, month]

    rows = _raw_query(f"""
        SELECT a.id, a.business_name,
               COUNT(DISTINCT t.id) as total_tasks,
               COUNT(DISTINCT CASE WHEN t.is_active=1 THEN t.id END) as active_tasks,
               COUNT(DISTINCT tc.user_id) as unique_participants,
               COUNT(CASE WHEN tc.status='approved' THEN 1 END) as completions,
               COALESCE(SUM(CASE WHEN tc.status='approved' THEN tc.points_awarded ELSE 0 END), 0) as points_issued,
               (SELECT COUNT(*) FROM redemptions r
                JOIN ally_products p ON p.id = r.reward_id
                WHERE p.ally_id = a.id {red_filter}) as redemptions
        FROM allies a
        LEFT JOIN tasks t ON t.ally_id = a.id
        LEFT JOIN task_completions tc ON tc.task_id = t.id {tc_filter}
        WHERE a.status = 'approved'
        GROUP BY a.id
        ORDER BY completions DESC
    """, params)
    return jsonify(rows)


@app.route("/api/admin/stats/by-task")
def admin_stats_by_task():
    """Estadísticas por tarea individual con filtros."""
    db.init_db()
    ally_id = request.args.get("ally_id", "")
    month = request.args.get("month", "")

    conditions = ["t.is_active = 1 OR t.is_active = 0"]
    params = []
    if ally_id and ally_id != "all":
        conditions.append("t.ally_id = ?")
        params.append(int(ally_id))

    sub_where = ""
    sub_params = []
    if month:
        sub_where = " AND strftime('%Y-%m', tc.submitted_at) = ?"
        sub_params.append(month)

    where_sql = " AND ".join(conditions)

    rows = _raw_query(f"""
        SELECT t.id, t.title, t.description, t.points_value, t.is_active,
               t.created_at,
               COALESCE(a.business_name, 'Sin empresa') as business_name,
               (SELECT COUNT(*) FROM task_completions tc WHERE tc.task_id = t.id AND tc.status='approved' {sub_where}) as approved,
               (SELECT COUNT(*) FROM task_completions tc WHERE tc.task_id = t.id AND tc.status='pending' {sub_where}) as pending,
               (SELECT COUNT(*) FROM task_completions tc WHERE tc.task_id = t.id AND tc.status='rejected' {sub_where}) as rejected,
               (SELECT COUNT(DISTINCT tc.user_id) FROM task_completions tc WHERE tc.task_id = t.id {sub_where}) as unique_users,
               (SELECT COALESCE(SUM(tc.points_awarded),0) FROM task_completions tc WHERE tc.task_id = t.id AND tc.status='approved' {sub_where}) as points_issued
        FROM tasks t
        LEFT JOIN allies a ON a.id = t.ally_id
        WHERE {where_sql}
        ORDER BY t.created_at DESC
    """, sub_params + sub_params + sub_params + sub_params + sub_params + params)
    return jsonify(rows)


@app.route("/api/admin/goals/<int:ally_id>")
def admin_list_goals(ally_id):
    db.init_db()
    rows = _raw_query(
        "SELECT * FROM ally_goals WHERE ally_id = ? ORDER BY id",
        (ally_id,)
    )
    return jsonify(rows)


@app.route("/api/admin/goals", methods=["POST"])
def admin_create_goal():
    db.init_db()
    data = request.json
    ally_id = data.get("ally_id")
    goal_type = data.get("goal_type", "").strip()
    if not ally_id or not goal_type:
        return jsonify({"ok": False, "error": "Empresa y tipo son obligatorios"}), 400
    conn = sqlite3.connect(DB_PATH)
    cur = conn.execute("""
        INSERT INTO ally_goals (ally_id, goal_type, target, current, period, notes)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (int(ally_id), goal_type, int(data.get("target", 0)),
          int(data.get("current", 0)), data.get("period", "monthly"),
          data.get("notes", "")))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/admin/goals/<int:gid>", methods=["PUT"])
def admin_update_goal(gid):
    db.init_db()
    data = request.json
    allowed = {"goal_type", "target", "current", "period", "notes"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"ok": True})
    set_clause = ", ".join(f"{k} = ?" for k in updates) + ", updated_at = datetime('now')"
    values = list(updates.values()) + [gid]
    conn = sqlite3.connect(DB_PATH)
    conn.execute(f"UPDATE ally_goals SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/goals/<int:gid>", methods=["DELETE"])
def admin_delete_goal(gid):
    db.init_db()
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM ally_goals WHERE id = ?", (gid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/export/monthly")
def admin_export_monthly():
    """Exporta un informe Excel del mes para una empresa o todas."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    db.init_db()
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    ally_id = request.args.get("ally_id", "all")

    wb = openpyxl.Workbook()

    # Sheet 1: Resumen por empresa
    ws1 = wb.active
    ws1.title = "Resumen Empresas"
    headers = ["Empresa", "Tareas activas", "Total tareas", "Participantes únicos",
               "Tareas completadas", "Puntos otorgados", "Canjes"]
    ws1.append(headers)
    for c in ws1[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1a2d47")
        c.alignment = Alignment(horizontal="center")

    where = "WHERE a.status='approved'"
    params = []
    if ally_id and ally_id != "all":
        where += " AND a.id = ?"
        params.append(int(ally_id))

    allies_rows = _raw_query(f"""
        SELECT a.id, a.business_name,
               COUNT(DISTINCT t.id) as total_tasks,
               COUNT(DISTINCT CASE WHEN t.is_active=1 THEN t.id END) as active_tasks,
               COUNT(DISTINCT CASE WHEN strftime('%Y-%m', tc.submitted_at) = ? THEN tc.user_id END) as unique_users,
               COUNT(CASE WHEN tc.status='approved' AND strftime('%Y-%m', tc.submitted_at) = ? THEN 1 END) as completions,
               COALESCE(SUM(CASE WHEN tc.status='approved' AND strftime('%Y-%m', tc.submitted_at) = ? THEN tc.points_awarded ELSE 0 END), 0) as points
        FROM allies a
        LEFT JOIN tasks t ON t.ally_id = a.id
        LEFT JOIN task_completions tc ON tc.task_id = t.id
        {where}
        GROUP BY a.id
        ORDER BY a.business_name
    """, [month, month, month] + params)

    for a in allies_rows:
        red_count = _raw_scalar("""
            SELECT COUNT(*) FROM redemptions r
            JOIN ally_products p ON p.id = r.reward_id
            WHERE p.ally_id = ? AND strftime('%Y-%m', r.redeemed_at) = ?
        """, (a["id"], month))
        ws1.append([
            a["business_name"], a["active_tasks"], a["total_tasks"],
            a["unique_users"], a["completions"], a["points"], red_count
        ])

    # Auto width
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 22

    # Sheet 2: Detalle de tareas
    ws2 = wb.create_sheet("Tareas detalle")
    headers2 = ["Empresa", "Tarea", "Puntos", "Aprobadas", "Pendientes", "Rechazadas", "Usuarios únicos", "Pts otorgados"]
    ws2.append(headers2)
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1a2d47")
        c.alignment = Alignment(horizontal="center")

    task_where = ""
    task_params = []
    if ally_id and ally_id != "all":
        task_where = "WHERE t.ally_id = ?"
        task_params.append(int(ally_id))

    tasks = _raw_query(f"""
        SELECT t.id, t.title, t.points_value,
               COALESCE(a.business_name, '-') as business_name,
               (SELECT COUNT(*) FROM task_completions tc WHERE tc.task_id = t.id AND tc.status='approved' AND strftime('%Y-%m', tc.submitted_at) = ?) as approved,
               (SELECT COUNT(*) FROM task_completions tc WHERE tc.task_id = t.id AND tc.status='pending' AND strftime('%Y-%m', tc.submitted_at) = ?) as pending,
               (SELECT COUNT(*) FROM task_completions tc WHERE tc.task_id = t.id AND tc.status='rejected' AND strftime('%Y-%m', tc.submitted_at) = ?) as rejected,
               (SELECT COUNT(DISTINCT tc.user_id) FROM task_completions tc WHERE tc.task_id = t.id AND strftime('%Y-%m', tc.submitted_at) = ?) as unique_users,
               (SELECT COALESCE(SUM(tc.points_awarded),0) FROM task_completions tc WHERE tc.task_id = t.id AND tc.status='approved' AND strftime('%Y-%m', tc.submitted_at) = ?) as points
        FROM tasks t
        LEFT JOIN allies a ON a.id = t.ally_id
        {task_where}
        ORDER BY a.business_name, t.id
    """, [month, month, month, month, month] + task_params)

    for t in tasks:
        ws2.append([t["business_name"], t["title"], t["points_value"],
                    t["approved"], t["pending"], t["rejected"],
                    t["unique_users"], t["points"]])

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 18

    # Sheet 3: Canjes del mes
    ws3 = wb.create_sheet("Canjes del mes")
    headers3 = ["Fecha", "Usuario", "Empresa", "Producto", "Puntos", "Codigo", "Estado"]
    ws3.append(headers3)
    for c in ws3[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1a2d47")
        c.alignment = Alignment(horizontal="center")

    red_where = "WHERE strftime('%Y-%m', r.redeemed_at) = ?"
    red_params = [month]
    if ally_id and ally_id != "all":
        red_where += " AND p.ally_id = ?"
        red_params.append(int(ally_id))

    reds = _raw_query(f"""
        SELECT r.redeemed_at, r.voucher_code, r.points_used, r.status,
               u.full_name, u.username,
               p.name as product_name,
               COALESCE(a.business_name, '-') as business_name
        FROM redemptions r
        JOIN ally_products p ON p.id = r.reward_id
        LEFT JOIN allies a ON a.id = p.ally_id
        JOIN users u ON u.user_id = r.user_id
        {red_where}
        ORDER BY r.redeemed_at DESC
    """, red_params)

    for r in reds:
        ws3.append([
            r["redeemed_at"][:16],
            r["full_name"] or r["username"] or "-",
            r["business_name"],
            r["product_name"],
            r["points_used"],
            r["voucher_code"],
            r["status"]
        ])

    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 20

    # Guardar a buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"informe_{month}"
    if ally_id and ally_id != "all":
        ally = db.get_ally(int(ally_id))
        if ally:
            safe = ally["business_name"].replace(" ", "_").replace("/", "_")
            filename += f"_{safe}"
    filename += ".xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ─── Static files ──────────────────────────────────────────────────────────

@app.route("/uploads/<path:filename>")
def serve_upload(filename):
    return send_from_directory(UPLOAD_DIR, filename)


@app.route("/api/aliados")
def listar_aliados():
    db.init_db()
    allies = db.list_allies(status="approved")
    result = []
    for a in allies:
        products = db.get_ally_products(a["id"])
        result.append({
            "id": a["id"], "business_name": a["business_name"],
            "location": a["location"], "city": a["city"],
            "description": a["description"],
            "photo": f"/uploads/{a['photo_path']}" if a["photo_path"] else None,
            "phone": a["phone"], "instagram": a["instagram"],
            "products": [{"name": p["name"], "description": p["description"],
                          "price": p["price"],
                          "photo": f"/uploads/{p['photo_path']}" if p["photo_path"] else None}
                         for p in products]
        })
    return jsonify(result)


if __name__ == "__main__":
    print("API corriendo en http://localhost:5055")
    app.run(host="0.0.0.0", port=5055, debug=False)
